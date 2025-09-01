#!/usr/bin/env python3
"""
Test script to simulate the exact Excel download flow and identify the issue.
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloader'))

from tiktok_downloader import TikTokDownloader
from pathlib import Path
import openpyxl

def test_excel_download_flow():
    """Test the exact Excel download flow."""
    
    print("🧪 Testing Excel Download Flow")
    print("=" * 50)
    
    # Initialize downloader
    downloader = TikTokDownloader(
        output_dir="downloads",
        quality="best",
        extract_audio=False,
        add_metadata=True
    )
    
    # Set custom base name
    downloader.custom_base_name = "flow_test"
    
    # Test URLs (these are example URLs - replace with real ones for testing)
    test_urls = [
        "https://www.tiktok.com/@user1/video/1234567890123456789",
        "https://www.tiktok.com/@user2/video/9876543210987654321",
        "https://www.tiktok.com/@user3/video/5556667778889990001"
    ]
    
    print(f"📊 Testing with {len(test_urls)} URLs")
    print(f"🎯 Custom base name: {downloader.custom_base_name}")
    print(f"📁 Output directory: {downloader.output_dir}")
    print(f"📄 Excel file: {downloader.excel_file}")
    
    # Reset video counter
    print("\n🔄 Resetting video counter...")
    downloader.reset_video_counter()
    print(f"📊 Video counter after reset: {downloader.video_counter}")
    
    # Simulate the download process for each URL
    for i, url in enumerate(test_urls, 1):
        print(f"\n🎬 Processing URL {i}: {url}")
        print("-" * 40)
        
        # Check if already downloaded BEFORE any download
        print("🔍 Checking if URL is already downloaded (BEFORE)...")
        already_downloaded = downloader.is_url_already_downloaded(url)
        print(f"📋 Already downloaded: {already_downloaded}")
        
        if already_downloaded:
            print("⏭️  Skipping already downloaded URL")
            continue
        
        print(f"📊 Video counter before processing: {downloader.video_counter}")
        
        # Simulate successful download by manually adding to Excel
        print("🔄 Simulating successful download...")
        
        # Create mock video info
        mock_info = {
            'id': f'video_{i}',
            'title': f'Test Video {i}',
            'description': f'Description for video {i}',
            'uploader': f'user{i}',
            'uploader_id': f'user{i}_id',
            'channel': f'channel{i}',
            'channel_id': f'channel{i}_id',
            'upload_date': '20250101',
            'duration': 30,
            'view_count': 1000,
            'like_count': 100,
            'comment_count': 50,
            'repost_count': 10,
            'ext': 'mp4',
            'filesize': 1024000,
            'width': 1920,
            'height': 1080,
            'format': 'mp4',
            'webpage_url': url
        }
        
        # Generate download path
        ext = mock_info.get('ext', 'mp4')
        current_counter = downloader.video_counter
        download_path = str(downloader.output_dir / f"{downloader.custom_base_name}__{current_counter}.{ext}")
        
        print(f"📝 Generated download path: {download_path}")
        
        # Add metadata to Excel (this simulates what happens after successful download)
        downloader._add_metadata_to_excel(mock_info, download_path)
        print("✅ Added metadata to Excel")
        
        # Increment video counter (simulate what _get_custom_filename does)
        downloader.video_counter += 1
        print(f"📊 Video counter after processing: {downloader.video_counter}")
        
        # Now check if the same URL is already downloaded (AFTER adding to Excel)
        print("🔍 Checking if URL is already downloaded (AFTER)...")
        already_downloaded_after = downloader.is_url_already_downloaded(url)
        print(f"📋 Already downloaded (AFTER): {already_downloaded_after}")
        
        # Check if the next URL would be incorrectly identified as already downloaded
        if i < len(test_urls):
            next_url = test_urls[i]
            print(f"🔍 Checking next URL: {next_url}")
            next_already_downloaded = downloader.is_url_already_downloaded(next_url)
            print(f"📋 Next URL already downloaded: {next_already_downloaded}")
            
            if next_already_downloaded:
                print("❌ PROBLEM: Next URL is incorrectly identified as already downloaded!")
                print("This is likely the cause of the download issue.")
    
    print(f"\n📊 Final video counter: {downloader.video_counter}")
    
    # Check Excel file content
    if downloader.excel_file.exists():
        print(f"\n📊 Excel file exists: {downloader.excel_file}")
        print("📋 Checking Excel file content...")
        
        try:
            wb = openpyxl.load_workbook(str(downloader.excel_file))
            ws = wb.active
            
            print(f"📊 Excel has {ws.max_row} rows (including header)")
            
            if ws.max_row > 1:
                print("📋 Data rows:")
                for row in range(2, ws.max_row + 1):
                    url = ws.cell(row=row, column=16).value  # Original URL column
                    path = ws.cell(row=row, column=len(downloader.headers)).value  # Download Path column
                    print(f"  Row {row}: URL={url}")
                    print(f"         Path={path}")
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
    else:
        print(f"\n📊 Excel file does not exist: {downloader.excel_file}")
    
    print("\n🎉 Excel download flow test completed!")

if __name__ == "__main__":
    test_excel_download_flow()
