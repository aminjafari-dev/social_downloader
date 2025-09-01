"""
Excel Loader Module

This module provides functionality for loading and reading Excel files.
It handles operations like reading column names, extracting URLs from specific columns,
and validating Excel file formats.

Usage:
    from utils.excel_loader import ExcelLoader
    
    loader = ExcelLoader()
    columns = loader.get_column_names("path/to/file.xlsx")
    urls = loader.extract_urls_from_column("path/to/file.xlsx", "URL Column")
"""

import os
import openpyxl
from pathlib import Path
from typing import List, Optional, Tuple
import logging

# Configure logging
logger = logging.getLogger(__name__)


class ExcelLoader:
    """
    A class to handle Excel file loading and reading operations.
    
    This class provides methods to read Excel files, extract column names,
    and retrieve data from specific columns. It's designed to work with
    TikTok video URL data but can be used for any Excel file.
    
    Attributes:
        None - This is a utility class with no instance state
    """
    
    def __init__(self):
        """Initialize the Excel loader."""
        pass
    
    def validate_excel_file(self, file_path: str) -> Tuple[bool, str]:
        """
        Validate if the provided file path is a valid Excel file.
        
        Args:
            file_path (str): Path to the Excel file to validate
            
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
                - is_valid: True if file is valid, False otherwise
                - error_message: Description of the error if invalid
        """
        if not file_path or not file_path.strip():
            return False, "No file path provided"
        
        file_path = file_path.strip()
        
        if not os.path.exists(file_path):
            return False, f"File does not exist: {file_path}"
        
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False, f"File is not an Excel file: {file_path}"
        
        try:
            # Try to open the file to verify it's a valid Excel file
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True, "File is valid"
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"
    
    def get_column_names(self, file_path: str) -> List[str]:
        """
        Extract column names from the first row of an Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            List[str]: List of column names from the first row
            
        Raises:
            ValueError: If file path is invalid or file cannot be read
        """
        # Validate file first
        is_valid, error_message = self.validate_excel_file(file_path)
        if not is_valid:
            raise ValueError(f"Cannot read Excel file: {error_message}")
        
        try:
            # Load workbook in read-only mode for better performance
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Extract column names from first row
            columns = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    columns.append(str(cell_value))
            
            wb.close()
            
            logger.info(f"Successfully loaded {len(columns)} columns from Excel file")
            return columns
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    def find_column_index(self, file_path: str, column_name: str) -> Optional[int]:
        """
        Find the index of a specific column by name.
        
        Args:
            file_path (str): Path to the Excel file
            column_name (str): Name of the column to find
            
        Returns:
            Optional[int]: Column index (1-based) if found, None otherwise
        """
        try:
            columns = self.get_column_names(file_path)
            
            # Find the column index (case-insensitive)
            for i, col in enumerate(columns, 1):
                if col.strip().lower() == column_name.strip().lower():
                    logger.info(f"Found column '{column_name}' at index {i}")
                    return i
            
            logger.warning(f"Column '{column_name}' not found in Excel file")
            return None
            
        except Exception as e:
            logger.error(f"Error finding column index: {str(e)}")
            return None
    
    def extract_data_from_column(self, file_path: str, column_name: str, 
                                start_row: int = 2) -> List[str]:
        """
        Extract all data from a specific column in the Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            column_name (str): Name of the column to extract data from
            start_row (int): Row to start extracting from (default: 2, skipping header)
            
        Returns:
            List[str]: List of values from the specified column
            
        Raises:
            ValueError: If file path is invalid or column not found
        """
        # Validate file first
        is_valid, error_message = self.validate_excel_file(file_path)
        if not is_valid:
            raise ValueError(f"Cannot read Excel file: {error_message}")
        
        # Find column index
        column_index = self.find_column_index(file_path, column_name)
        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in Excel file")
        
        try:
            # Load workbook in read-only mode
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Extract data from the specified column
            data = []
            for row in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=column_index).value
                if cell_value:
                    data.append(str(cell_value).strip())
            
            wb.close()
            
            logger.info(f"Successfully extracted {len(data)} values from column '{column_name}'")
            return data
            
        except Exception as e:
            logger.error(f"Error extracting data from column: {str(e)}")
            raise ValueError(f"Failed to extract data from column: {str(e)}")
    
    def extract_urls_from_column(self, file_path: str, column_name: str, 
                                url_validator=None) -> List[str]:
        """
        Extract URLs from a specific column and optionally validate them.
        
        Args:
            file_path (str): Path to the Excel file
            column_name (str): Name of the column containing URLs
            url_validator (callable, optional): Function to validate URLs
                Should take a URL string and return True if valid, False otherwise
            
        Returns:
            List[str]: List of valid URLs from the specified column
        """
        try:
            # Extract all data from the column
            all_data = self.extract_data_from_column(file_path, column_name)
            
            if url_validator is None:
                # If no validator provided, return all non-empty values
                valid_urls = [url for url in all_data if url.strip()]
            else:
                # Validate URLs using the provided validator
                valid_urls = []
                for url in all_data:
                    if url.strip() and url_validator(url.strip()):
                        valid_urls.append(url.strip())
                    elif url.strip():
                        logger.warning(f"Invalid URL found: {url}")
            
            logger.info(f"Extracted {len(valid_urls)} valid URLs from column '{column_name}'")
            return valid_urls
            
        except Exception as e:
            logger.error(f"Error extracting URLs from column: {str(e)}")
            return []
    
    def get_excel_preview(self, file_path: str, column_name: str, 
                         max_preview: int = 5) -> str:
        """
        Generate a preview of data from a specific column.
        
        Args:
            file_path (str): Path to the Excel file
            column_name (str): Name of the column to preview
            max_preview (int): Maximum number of items to show in preview
            
        Returns:
            str: Formatted preview string
        """
        try:
            data = self.extract_data_from_column(file_path, column_name)
            
            preview_text = f"Total items in column '{column_name}': {len(data)}\n\n"
            preview_text += f"First {min(max_preview, len(data))} items:\n"
            
            for i, item in enumerate(data[:max_preview], 1):
                # Truncate long items
                display_item = item[:100] + "..." if len(item) > 100 else item
                preview_text += f"{i}. {display_item}\n"
            
            if len(data) > max_preview:
                preview_text += f"\n... and {len(data) - max_preview} more items"
            
            return preview_text
            
        except Exception as e:
            return f"Error generating preview: {str(e)}"
    
    def get_excel_info(self, file_path: str) -> dict:
        """
        Get basic information about an Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            dict: Dictionary containing file information
        """
        try:
            # Validate file first
            is_valid, error_message = self.validate_excel_file(file_path)
            if not is_valid:
                return {"error": error_message}
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get basic info
            info = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "total_rows": ws.max_row,
                "total_columns": ws.max_column,
                "data_rows": max(0, ws.max_row - 1),  # Exclude header row
                "column_names": self.get_column_names(file_path)
            }
            
            wb.close()
            return info
            
        except Exception as e:
            return {"error": f"Failed to get file info: {str(e)}"}


# Convenience functions for backward compatibility
def load_excel_columns(file_path: str) -> List[str]:
    """
    Convenience function to load column names from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        List[str]: List of column names
    """
    loader = ExcelLoader()
    return loader.get_column_names(file_path)


def extract_urls_from_excel(file_path: str, column_name: str, 
                           url_validator=None) -> List[str]:
    """
    Convenience function to extract URLs from an Excel column.
    
    Args:
        file_path (str): Path to the Excel file
        column_name (str): Name of the column containing URLs
        url_validator (callable, optional): Function to validate URLs
        
    Returns:
        List[str]: List of valid URLs
    """
    loader = ExcelLoader()
    return loader.extract_urls_from_column(file_path, column_name, url_validator)


def validate_excel_file(file_path: str) -> Tuple[bool, str]:
    """
    Convenience function to validate an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        Tuple[bool, str]: (is_valid, error_message)
    """
    loader = ExcelLoader()
    return loader.validate_excel_file(file_path)
