"""
TikTok-Specific Video Downloader Module

This module extends the core VideoDownloader to provide TikTok-specific functionality.
It includes TikTok URL validation, metadata extraction, and specialized handling.

Usage:
    from core.tiktok_downloader import TikTokDownloader
    
    downloader = TikTokDownloader(output_dir="downloads", quality="best")
    success = downloader.download_video("https://www.tiktok.com/@user/video/1234567890")
"""

import logging
from pathlib import Path
from typing import Optional, Dict, Any, List
from datetime import datetime

from .video_downloader import VideoDownloader

# Configure logging
logger = logging.getLogger(__name__)


class TikTokDownloader(VideoDownloader):
    """
    TikTok-specific video downloader that extends the core VideoDownloader.
    
    This class adds TikTok-specific functionality:
    - TikTok URL validation
    - TikTok metadata extraction
    - TikTok-specific file naming
    
    Attributes:
        Inherits all attributes from VideoDownloader
        tiktok_domains (List[str]): List of valid TikTok domains
    """
    
    def __init__(self, output_dir: str = "downloads", quality: str = "best", 
                 extract_audio: bool = False, add_metadata: bool = True,
                 custom_base_name: str = None):
        """
        Initialize the TikTok downloader with specified options.
        
        Args:
            output_dir (str): Directory to save downloaded videos (default: "downloads")
            quality (str): Preferred video quality (default: "best")
            extract_audio (bool): Extract audio only if True (default: False)
            add_metadata (bool): Add metadata to downloaded files (default: True)
            custom_base_name (str): Custom base name for video files (default: None)
        """
        super().__init__(output_dir, quality, extract_audio, add_metadata, custom_base_name)
        
        # TikTok-specific domains
        self.tiktok_domains = [
            'tiktok.com',
            'vm.tiktok.com',
            'vt.tiktok.com',
            'www.tiktok.com'
        ]
        
        logger.info("TikTokDownloader initialized with TikTok-specific functionality")
    
    def validate_url(self, url: str) -> bool:
        """
        Validate if the provided URL is a valid TikTok URL.
        
        Args:
            url (str): The URL to validate
            
        Returns:
            bool: True if URL is a valid TikTok URL, False otherwise
        """
        if not super().validate_url(url):
            return False
        
        try:
            import urllib.parse
            parsed = urllib.parse.urlparse(url.strip())
            return any(domain in parsed.netloc for domain in self.tiktok_domains)
        except Exception:
            return False
    
    def extract_tiktok_metadata(self, info: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract TikTok-specific metadata from video info.
        
        Args:
            info (Dict[str, Any]): Video information dictionary from yt-dlp
            
        Returns:
            Dict[str, Any]: Extracted TikTok metadata
        """
        metadata = {
            'video_id': info.get('id', ''),
            'title': info.get('title', ''),
            'description': info.get('description', ''),
            'uploader': info.get('uploader', ''),
            'uploader_id': info.get('uploader_id', ''),
            'channel': info.get('channel', ''),
            'channel_id': info.get('channel_id', ''),
            'upload_date': info.get('upload_date', ''),
            'duration': info.get('duration', 0),
            'view_count': info.get('view_count', 0),
            'like_count': info.get('like_count', 0),
            'comment_count': info.get('comment_count', 0),
            'repost_count': info.get('repost_count', 0),
            'hashtags': self._extract_hashtags(info.get('description', '')),
            'original_url': info.get('webpage_url', ''),
            'thumbnail_url': info.get('thumbnail', ''),
            'video_quality': self._get_video_quality(info),
            'file_size': self._get_file_size(info),
            'resolution': self._get_resolution(info),
            'format': info.get('ext', ''),
            'download_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        logger.debug(f"Extracted TikTok metadata: {metadata}")
        return metadata
    
    def _extract_hashtags(self, description: str) -> str:
        """
        Extract hashtags from video description.
        
        Args:
            description (str): Video description text
            
        Returns:
            str: Comma-separated hashtags
        """
        if not description:
            return ""
        
        hashtags = []
        words = description.split()
        for word in words:
            if word.startswith('#'):
                hashtags.append(word)
        
        return ', '.join(hashtags)
    
    def _get_video_quality(self, info: Dict[str, Any]) -> str:
        """
        Get video quality information.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            
        Returns:
            str: Video quality string
        """
        format_info = info.get('format', '')
        if isinstance(format_info, str):
            return format_info
        else:
            return format_info.get('format_note', '') or format_info.get('format', '')
    
    def _get_file_size(self, info: Dict[str, Any]) -> int:
        """
        Get file size information.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            
        Returns:
            int: File size in bytes
        """
        filesize = info.get('filesize', 0)
        if not filesize and 'format' in info:
            filesize = info['format'].get('filesize', 0)
        return filesize
    
    def _get_resolution(self, info: Dict[str, Any]) -> str:
        """
        Get video resolution information.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            
        Returns:
            str: Resolution string (e.g., "1920x1080")
        """
        width = info.get('width', 0)
        height = info.get('height', 0)
        return f"{width}x{height}" if width and height else ""
    
    def format_upload_date(self, upload_date: str) -> str:
        """
        Format upload date from TikTok format to readable format.
        
        Args:
            upload_date (str): Upload date in TikTok format (YYYYMMDD)
            
        Returns:
            str: Formatted date string (YYYY-MM-DD)
        """
        if not upload_date or len(upload_date) < 8:
            return ""
        
        try:
            return f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:8]}"
        except:
            return upload_date
    
    def get_download_path(self, info: Dict[str, Any]) -> str:
        """
        Get the path where the video was downloaded.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            
        Returns:
            str: Download path string
        """
        ext = info.get('ext', 'mp4')
        
        if self.custom_base_name:
            # Use custom naming pattern
            current_counter = self.video_counter - 1  # Adjust for the increment in _get_custom_filename
            possible_filename = f"{self.custom_base_name}__{current_counter}.{ext}"
            download_path = str(self.output_dir / possible_filename)
            
            # Check if file exists
            if not Path(download_path).exists():
                # Try to find any file with the custom naming pattern
                for file in self.output_dir.glob(f"{self.custom_base_name}__*.{ext}"):
                    if f"__{current_counter}." in file.name:
                        download_path = str(file)
                        break
        else:
            # Use original title-based naming
            if info.get('title'):
                title = info.get('title', '')
                possible_filename = f"{title}.{ext}"
                download_path = str(self.output_dir / possible_filename)
                
                # Check if file exists
                if not Path(download_path).exists():
                    # Try to find any file with similar name
                    for file in self.output_dir.glob(f"*.{ext}"):
                        if title.lower() in file.name.lower():
                            download_path = str(file)
                            break
        
        return download_path
    
    def is_url_already_downloaded(self, url: str, excel_file: Path = None) -> bool:
        """
        Check if a TikTok URL has already been downloaded by checking an Excel file.
        
        Args:
            url (str): The TikTok URL to check
            excel_file (Path): Path to Excel file to check (optional)
            
        Returns:
            bool: True if URL already exists in Excel, False otherwise
        """
        if not excel_file or not excel_file.exists():
            return False
        
        try:
            # Load existing Excel file
            import openpyxl
            wb = openpyxl.load_workbook(str(excel_file))
            ws = wb.active
            
            # Check if URL exists in the "Original URL" column (column 16)
            for row in range(2, ws.max_row + 1):
                existing_url = ws.cell(row=row, column=16).value
                if existing_url == url:
                    wb.close()
                    return True
            
            wb.close()
            return False
            
        except Exception as e:
            logger.warning(f"Could not check Excel file for existing URL: {e}")
            return False
