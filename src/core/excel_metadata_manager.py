"""
Excel Metadata Manager Module

This module handles all Excel-related operations for storing video metadata.
It provides functionality to create, update, and manage Excel files containing video information.

Usage:
    from core.excel_metadata_manager import ExcelMetadataManager
    
    manager = ExcelMetadataManager(output_dir="downloads", filename="videos_metadata.xlsx")
    manager.add_video_metadata(video_info, download_path)
    manager.save_excel_file()
"""

import os
import logging
import openpyxl
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logger = logging.getLogger(__name__)


class ExcelMetadataManager:
    """
    Manages Excel files for storing video metadata.
    
    This class handles:
    - Creating and configuring Excel workbooks
    - Adding video metadata to Excel sheets
    - Formatting and styling Excel data
    - Saving and managing Excel files
    
    Attributes:
        output_dir (Path): Directory where Excel files will be saved
        excel_file (Path): Path to the Excel file
        workbook (openpyxl.Workbook): Excel workbook object
        worksheet (openpyxl.worksheet.worksheet.Worksheet): Active worksheet
        headers (List[str]): List of column headers for metadata
    """
    
    def __init__(self, output_dir: str = "downloads", filename: str = None):
        """
        Initialize the Excel metadata manager.
        
        Args:
            output_dir (str): Directory to save Excel files (default: "downloads")
            filename (str): Excel filename (default: auto-generated with timestamp)
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Set Excel file path
        if filename is None:
            # Use a consistent filename without timestamp to maintain data across sessions
            self.excel_file = self.output_dir / "videos_metadata.xlsx"
        else:
            self.excel_file = self.output_dir / filename
        
        # Define headers for Excel
        self.headers = [
            'Video ID', 'Title', 'Description', 'Uploader', 'Uploader ID', 
            'Channel', 'Channel ID', 'Upload Date', 'Duration (seconds)', 
            'Duration (formatted)', 'View Count', 'Like Count', 'Comment Count', 
            'Repost Count', 'Hashtags', 'Original URL', 'Thumbnail URL',
            'Video Quality', 'File Size (bytes)', 'Resolution', 'Format',
            'Download Date', 'Download Path'
        ]
        
        # Try to load existing Excel file, create new one if it doesn't exist
        if self.excel_file.exists():
            logger.info(f"Loading existing Excel file: {self.excel_file}")
            self._load_existing_excel()
        else:
            logger.info(f"Creating new Excel file: {self.excel_file}")
            self._create_new_excel()
        
        logger.info(f"ExcelMetadataManager initialized with file: {self.excel_file}")
    
    def _create_new_excel(self):
        """Create a new Excel workbook with headers."""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Video Metadata"
        self._setup_excel_headers()
        logger.info("Created new Excel workbook with headers")
    
    def _load_existing_excel(self):
        """Load an existing Excel file and validate its structure."""
        try:
            self.workbook = openpyxl.load_workbook(str(self.excel_file))
            self.worksheet = self.workbook.active
            
            # Validate that the file has the expected headers
            if self.worksheet.max_row > 0:
                existing_headers = []
                for col in range(1, self.worksheet.max_column + 1):
                    header_value = self.worksheet.cell(row=1, column=col).value
                    existing_headers.append(header_value)
                
                # Check if headers match our expected format
                if len(existing_headers) >= len(self.headers):
                    logger.info(f"Loaded existing Excel file with {self.worksheet.max_row - 1} data rows")
                else:
                    logger.warning("Existing Excel file has different structure, creating new one")
                    self._create_new_excel()
            else:
                logger.info("Existing Excel file is empty, setting up headers")
                self._setup_excel_headers()
                
        except Exception as e:
            logger.error(f"Error loading existing Excel file: {e}")
            logger.info("Creating new Excel file due to loading error")
            self._create_new_excel()
    
    def _setup_excel_headers(self):
        """Setup Excel worksheet with headers and formatting."""
        # Add headers
        for col, header in enumerate(self.headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col in range(1, len(self.headers) + 1):
            column_letter = get_column_letter(col)
            self.worksheet.column_dimensions[column_letter].width = 15
        
        logger.debug("Excel headers configured and formatted")
    
    def _format_duration(self, duration: int) -> str:
        """
        Format duration from seconds to MM:SS format.
        
        Args:
            duration (int): Duration in seconds
            
        Returns:
            str: Formatted duration string
        """
        if not duration:
            return ""
        
        minutes = duration // 60
        seconds = duration % 60
        return f"{minutes:02d}:{seconds:02d}"
    
    def _extract_hashtags(self, description: str) -> str:
        """
        Extract hashtags from video description.
        
        Args:
            description (str): Video description text
            
        Returns:
            str: Comma-separated hashtags
        """
        if not description:
            return ""
        
        hashtags = []
        words = description.split()
        for word in words:
            if word.startswith('#'):
                hashtags.append(word)
        
        return ', '.join(hashtags)
    
    def _extract_metadata_for_excel(self, info: Dict[str, Any], download_path: str = "") -> List[Any]:
        """
        Extract metadata from video info for Excel export.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            download_path (str): Path where video was downloaded
            
        Returns:
            List[Any]: List of metadata values for Excel row
        """
        # Extract hashtags from description
        description = info.get('description', '')
        hashtags = self._extract_hashtags(description)
        
        # Format duration
        duration = info.get('duration', 0)
        duration_formatted = self._format_duration(duration)
        
        # Get video format info
        format_info = info.get('format', '')
        if isinstance(format_info, str):
            video_quality = format_info
        else:
            video_quality = format_info.get('format_note', '') or format_info.get('format', '')
        
        # Get file size
        filesize = info.get('filesize', 0)
        if not filesize and 'format' in info:
            filesize = info['format'].get('filesize', 0)
        
        # Get resolution
        width = info.get('width', 0)
        height = info.get('height', 0)
        resolution = f"{width}x{height}" if width and height else ""
        
        # Get thumbnail URL
        thumbnail = info.get('thumbnail', '')
        
        # Format upload date
        upload_date = info.get('upload_date', '')
        if upload_date:
            try:
                formatted_date = f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:8]}"
            except:
                formatted_date = upload_date
        else:
            formatted_date = ""
        
        # Current download date
        download_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        return [
            info.get('id', ''),
            info.get('title', ''),
            description,
            info.get('uploader', ''),
            info.get('uploader_id', ''),
            info.get('channel', ''),
            info.get('channel_id', ''),
            formatted_date,
            duration,
            duration_formatted,
            info.get('view_count', 0),
            info.get('like_count', 0),
            info.get('comment_count', 0),
            info.get('repost_count', 0),
            hashtags,
            info.get('webpage_url', ''),
            thumbnail,
            video_quality,
            filesize,
            resolution,
            info.get('ext', ''),
            download_date,
            download_path
        ]
    
    def add_video_metadata(self, info: Dict[str, Any], download_path: str = ""):
        """
        Add video metadata to Excel worksheet.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            download_path (str): Path where video was downloaded
        """
        try:
            # Check if video already exists in Excel to avoid duplicates
            video_id = info.get('id', '')
            original_url = info.get('webpage_url', '')
            
            if self._video_exists(video_id, original_url):
                logger.info(f"Video already exists in Excel, skipping: {info.get('title', 'Unknown')}")
                return
            
            # Get next row number
            next_row = self.worksheet.max_row + 1
            
            # Extract metadata
            metadata = self._extract_metadata_for_excel(info, download_path)
            
            # Add data to worksheet
            for col, value in enumerate(metadata, 1):
                cell = self.worksheet.cell(row=next_row, column=col, value=value)
                
                # Format numbers
                if col in [11, 12, 13, 14, 19]:  # View count, like count, comment count, repost count, file size
                    if value and value != 0:
                        cell.number_format = '#,##0'
            
            # Auto-adjust column widths
            for col in range(1, len(self.headers) + 1):
                column_letter = get_column_letter(col)
                current_width = self.worksheet.column_dimensions[column_letter].width
                cell_value = self.worksheet.cell(row=next_row, column=col).value
                if cell_value:
                    content_length = len(str(cell_value))
                    if content_length > current_width:
                        self.worksheet.column_dimensions[column_letter].width = min(content_length + 2, 50)
            
            logger.info(f"Added metadata to Excel: {info.get('title', 'Unknown')}")
            
        except Exception as e:
            logger.error(f"Error adding metadata to Excel: {e}")
    
    def _video_exists(self, video_id: str, original_url: str) -> bool:
        """
        Check if a video already exists in the Excel file.
        
        Args:
            video_id (str): Video ID to check
            original_url (str): Original URL to check
            
        Returns:
            bool: True if video exists, False otherwise
        """
        if not video_id and not original_url:
            return False
        
        # Find the column indices for Video ID and Original URL
        video_id_col = None
        url_col = None
        
        for col in range(1, self.worksheet.max_column + 1):
            header_value = self.worksheet.cell(row=1, column=col).value
            if header_value == 'Video ID':
                video_id_col = col
            elif header_value == 'Original URL':
                url_col = col
        
        # Check existing rows for duplicates
        for row in range(2, self.worksheet.max_row + 1):
            existing_video_id = self.worksheet.cell(row=row, column=video_id_col).value if video_id_col else None
            existing_url = self.worksheet.cell(row=row, column=url_col).value if url_col else None
            
            # Check if video ID matches
            if video_id and existing_video_id and str(video_id) == str(existing_video_id):
                return True
            
            # Check if URL matches
            if original_url and existing_url and str(original_url) == str(existing_url):
                return True
        
        return False
    
    def save_excel_file(self):
        """Save the Excel file with all collected metadata."""
        try:
            self.workbook.save(str(self.excel_file))
            logger.info(f"Excel file saved: {self.excel_file}")
            return True
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            return False
    
    def load_existing_excel(self, file_path: str) -> bool:
        """
        Load an existing Excel file to continue adding metadata.
        
        Args:
            file_path (str): Path to existing Excel file
            
        Returns:
            bool: True if loaded successfully, False otherwise
        """
        try:
            if not os.path.exists(file_path):
                logger.warning(f"Excel file does not exist: {file_path}")
                return False
            
            # Update the Excel file path
            self.excel_file = Path(file_path)
            
            # Load existing workbook
            self.workbook = openpyxl.load_workbook(file_path)
            self.worksheet = self.workbook.active
            
            # Validate the structure
            if self.worksheet.max_row > 0:
                existing_headers = []
                for col in range(1, self.worksheet.max_column + 1):
                    header_value = self.worksheet.cell(row=1, column=col).value
                    existing_headers.append(header_value)
                
                logger.info(f"Loaded existing Excel file: {file_path} with {self.worksheet.max_row - 1} data rows")
            else:
                logger.info(f"Loaded existing Excel file: {file_path} (empty)")
            
            return True
            
        except Exception as e:
            logger.error(f"Error loading existing Excel file: {e}")
            return False
    
    def process_existing_downloads(self, output_dir: str, custom_base_name: str = None) -> int:
        """
        Process existing downloaded videos and add their metadata to Excel.
        
        Args:
            output_dir (str): Directory containing downloaded videos
            custom_base_name (str): Custom base name for video files (optional)
            
        Returns:
            int: Number of videos processed
        """
        output_path = Path(output_dir)
        if not output_path.exists():
            logger.warning(f"Output directory does not exist: {output_dir}")
            return 0
        
        # Find all .info.json files in the output directory
        info_files = list(output_path.glob("*.info.json"))
        
        if not info_files:
            logger.info(f"No existing .info.json files found in {output_dir}")
            return 0
        
        logger.info(f"Found {len(info_files)} existing video metadata files")
        
        processed_count = 0
        for info_file in info_files:
            try:
                # Read the JSON file
                import json
                with open(info_file, 'r', encoding='utf-8') as f:
                    info = json.load(f)
                
                # Find corresponding video file
                video_path = ""
                title = info.get('title', '')
                ext = info.get('ext', 'mp4')
                
                if custom_base_name:
                    # Look for video file with custom naming pattern
                    for video_file in output_path.glob(f"{custom_base_name}__*.{ext}"):
                        # Check if this file corresponds to the current info file
                        info_base = info_file.stem.replace('.info', '')
                        if info_base in video_file.stem:
                            video_path = str(video_file)
                            break
                else:
                    # Look for video file with similar name
                    for video_file in output_path.glob(f"*.{ext}"):
                        if title.lower() in video_file.name.lower():
                            video_path = str(video_file)
                            break
                
                # Add to Excel
                self.add_video_metadata(info, video_path)
                processed_count += 1
                
                logger.info(f"Processed: {title[:50]}...")
                
            except Exception as e:
                logger.error(f"Error processing {info_file}: {e}")
        
        logger.info(f"Successfully processed {processed_count} videos")
        return processed_count
    
    def get_excel_info(self) -> dict:
        """
        Get basic information about the current Excel file.
        
        Returns:
            dict: Dictionary containing file information
        """
        try:
            info = {
                "file_path": str(self.excel_file),
                "file_name": self.excel_file.name,
                "total_rows": self.worksheet.max_row,
                "total_columns": self.worksheet.max_column,
                "data_rows": max(0, self.worksheet.max_row - 1),  # Exclude header row
                "column_names": self.headers,
                "file_exists": self.excel_file.exists(),
                "is_new_file": not self.excel_file.exists() or self.worksheet.max_row <= 1
            }
            
            return info
            
        except Exception as e:
            logger.error(f"Failed to get file info: {str(e)}")
            return {"error": f"Failed to get file info: {str(e)}"}
    
    def clear_data(self):
        """Clear all data from the worksheet, keeping only headers."""
        try:
            # Delete all rows except the first (header) row
            for row in range(self.worksheet.max_row, 1, -1):
                self.worksheet.delete_rows(row)
            
            logger.info("Cleared all data from Excel worksheet")
            
        except Exception as e:
            logger.error(f"Error clearing data: {e}")
    
    def close_workbook(self):
        """Close the Excel workbook to free up resources."""
        try:
            self.workbook.close()
            logger.info("Excel workbook closed")
        except Exception as e:
            logger.error(f"Error closing workbook: {e}")
    
    def get_file_status(self) -> str:
        """
        Get a human-readable status of the Excel file.
        
        Returns:
            str: Status message describing the current state
        """
        try:
            if not self.excel_file.exists():
                return "New file will be created"
            
            data_rows = max(0, self.worksheet.max_row - 1)
            if data_rows == 0:
                return "Empty file - ready for new data"
            else:
                return f"Existing file with {data_rows} videos - will append new data"
                
        except Exception as e:
            return f"Error checking file status: {str(e)}"
