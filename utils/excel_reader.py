"""
Excel Reader Module

This module provides functionality for reading, displaying, and analyzing Excel files.
It includes features for showing data in a readable format, finding the most recent
Excel files, and extracting specific information from Excel data.

Usage:
    from utils.excel_reader import ExcelReader
    
    reader = ExcelReader()
    reader.show_excel_data("downloads/tiktok_videos_metadata.xlsx")
    latest_file = reader.find_latest_excel_file("downloads")
"""

import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging
from datetime import datetime

# Configure logging
logger = logging.getLogger(__name__)


class ExcelReader:
    """
    A class to handle Excel file reading and display operations.
    
    This class provides methods to read Excel files, display data in a readable format,
    find the most recent Excel files, and extract specific information from Excel data.
    
    Attributes:
        None - This is a utility class with no instance state
    """
    
    def __init__(self):
        """Initialize the Excel reader."""
        pass
    
    def find_latest_excel_file(self, directory: str, pattern: str = "*.xlsx") -> Optional[Path]:
        """
        Find the most recent Excel file in a directory.
        
        Args:
            directory (str): Directory to search in
            pattern (str): File pattern to match (default: "*.xlsx")
            
        Returns:
            Optional[Path]: Path to the most recent Excel file, or None if not found
        """
        try:
            dir_path = Path(directory)
            if not dir_path.exists():
                logger.warning(f"Directory does not exist: {directory}")
                return None
            
            # Find all Excel files matching the pattern
            excel_files = list(dir_path.glob(pattern))
            
            if not excel_files:
                logger.info(f"No Excel files found in {directory} matching pattern '{pattern}'")
                return None
            
            # Get the most recent file based on modification time
            latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
            
            logger.info(f"Found latest Excel file: {latest_file}")
            return latest_file
            
        except Exception as e:
            logger.error(f"Error finding latest Excel file: {str(e)}")
            return None
    
    def load_excel_workbook(self, file_path: str) -> Optional[openpyxl.Workbook]:
        """
        Load an Excel workbook from file.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            Optional[openpyxl.Workbook]: Loaded workbook or None if failed
        """
        try:
            if not Path(file_path).exists():
                logger.error(f"Excel file does not exist: {file_path}")
                return None
            
            workbook = openpyxl.load_workbook(file_path)
            logger.info(f"Successfully loaded Excel file: {file_path}")
            return workbook
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            return None
    
    def get_excel_headers(self, workbook: openpyxl.Workbook) -> List[str]:
        """
        Extract headers from the first row of an Excel workbook.
        
        Args:
            workbook (openpyxl.Workbook): Loaded Excel workbook
            
        Returns:
            List[str]: List of column headers
        """
        try:
            worksheet = workbook.active
            headers = []
            
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Column_{col}")
            
            logger.info(f"Extracted {len(headers)} headers from Excel file")
            return headers
            
        except Exception as e:
            logger.error(f"Error extracting headers: {str(e)}")
            return []
    
    def get_excel_data_summary(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Get a summary of Excel data including row count, column count, etc.
        
        Args:
            workbook (openpyxl.Workbook): Loaded Excel workbook
            
        Returns:
            Dict[str, Any]: Summary information about the Excel data
        """
        try:
            worksheet = workbook.active
            headers = self.get_excel_headers(workbook)
            
            summary = {
                "total_rows": worksheet.max_row,
                "total_columns": worksheet.max_column,
                "data_rows": max(0, worksheet.max_row - 1),  # Exclude header row
                "headers": headers,
                "header_count": len(headers)
            }
            
            logger.info(f"Generated Excel data summary: {summary['data_rows']} data rows, {summary['header_count']} columns")
            return summary
            
        except Exception as e:
            logger.error(f"Error generating data summary: {str(e)}")
            return {}
    
    def format_excel_data_display(self, workbook: openpyxl.Workbook, 
                                 max_videos: int = 10, 
                                 key_fields: List[tuple] = None) -> str:
        """
        Format Excel data for display in a readable format.
        
        Args:
            workbook (openpyxl.Workbook): Loaded Excel workbook
            max_videos (int): Maximum number of videos to display
            key_fields (List[tuple]): List of (field_name, column_index) tuples for key fields
            
        Returns:
            str: Formatted display string
        """
        try:
            worksheet = workbook.active
            headers = self.get_excel_headers(workbook)
            summary = self.get_excel_data_summary(workbook)
            
            # Default key fields if not provided
            if key_fields is None:
                key_fields = [
                    ('Title', 2),
                    ('Uploader', 4),
                    ('Duration', 10),
                    ('View Count', 11),
                    ('Like Count', 12),
                    ('Comment Count', 13),
                    ('Repost Count', 14),
                    ('Hashtags', 15),
                    ('Upload Date', 8),
                    ('Video ID', 1)
                ]
            
            # Build display string
            display_text = f"📊 Excel Data Summary\n"
            display_text += "=" * 80 + "\n"
            display_text += f"📋 Found {summary['header_count']} columns:\n"
            
            for i, header in enumerate(headers, 1):
                display_text += f"   {i:2d}. {header}\n"
            
            display_text += f"\n📊 Found {summary['data_rows']} videos:\n"
            display_text += "=" * 80 + "\n"
            
            # Display data for each video (limited by max_videos)
            end_row = min(worksheet.max_row, max_videos + 1)  # +1 for header row
            
            for row in range(2, end_row + 1):
                display_text += f"\n🎬 Video {row - 1}:\n"
                display_text += "-" * 40 + "\n"
                
                # Display key information
                for field_name, col in key_fields:
                    if col <= len(headers):
                        value = worksheet.cell(row=row, column=col).value
                        if value:
                            # Truncate long values
                            if isinstance(value, str) and len(value) > 50:
                                value = value[:47] + "..."
                            display_text += f"   {field_name}: {value}\n"
                
                # Show full description if available
                description = worksheet.cell(row=row, column=3).value
                if description:
                    display_text += f"   Description: {description[:100]}...\n"
            
            if summary['data_rows'] > max_videos:
                display_text += f"\n... and {summary['data_rows'] - max_videos} more videos"
            
            display_text += "\n" + "=" * 80 + "\n"
            display_text += f"✅ Excel file contains comprehensive metadata for {summary['data_rows']} videos"
            
            return display_text
            
        except Exception as e:
            logger.error(f"Error formatting Excel data display: {str(e)}")
            return f"Error formatting Excel data: {str(e)}"
    
    def show_excel_data(self, file_path: str = None, directory: str = "downloads", 
                       pattern: str = "*.xlsx") -> str:
        """
        Display Excel data in a readable format.
        
        Args:
            file_path (str, optional): Specific Excel file path
            directory (str): Directory to search for Excel files if file_path not provided
            pattern (str): File pattern to match when searching directory
            
        Returns:
            str: Formatted display string
        """
        try:
            # Determine which file to use
            if file_path:
                target_file = Path(file_path)
            else:
                target_file = self.find_latest_excel_file(directory, pattern)
            
            if not target_file or not target_file.exists():
                return "❌ No Excel files found!"
            
            print(f"📊 Reading Excel file: {target_file}")
            
            # Load workbook
            workbook = self.load_excel_workbook(str(target_file))
            if not workbook:
                return "❌ Failed to load Excel file"
            
            # Format and return display
            display_text = self.format_excel_data_display(workbook)
            display_text += f"\n📁 File location: {target_file}"
            
            workbook.close()
            return display_text
            
        except Exception as e:
            logger.error(f"Error showing Excel data: {str(e)}")
            return f"❌ Error displaying Excel data: {str(e)}"
    
    def extract_specific_data(self, workbook: openpyxl.Workbook, 
                            column_name: str, row_range: tuple = None) -> List[Any]:
        """
        Extract specific data from a column in the Excel file.
        
        Args:
            workbook (openpyxl.Workbook): Loaded Excel workbook
            column_name (str): Name of the column to extract data from
            row_range (tuple, optional): (start_row, end_row) range to extract from
            
        Returns:
            List[Any]: List of values from the specified column
        """
        try:
            worksheet = workbook.active
            headers = self.get_excel_headers(workbook)
            
            # Find column index
            column_index = None
            for i, header in enumerate(headers, 1):
                if header.lower() == column_name.lower():
                    column_index = i
                    break
            
            if column_index is None:
                logger.warning(f"Column '{column_name}' not found in Excel file")
                return []
            
            # Determine row range
            if row_range:
                start_row, end_row = row_range
            else:
                start_row = 2  # Skip header
                end_row = worksheet.max_row
            
            # Extract data
            data = []
            for row in range(start_row, end_row + 1):
                cell_value = worksheet.cell(row=row, column=column_index).value
                if cell_value is not None:
                    data.append(cell_value)
            
            logger.info(f"Extracted {len(data)} values from column '{column_name}'")
            return data
            
        except Exception as e:
            logger.error(f"Error extracting specific data: {str(e)}")
            return []
    
    def get_excel_statistics(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Generate statistics from Excel data.
        
        Args:
            workbook (openpyxl.Workbook): Loaded Excel workbook
            
        Returns:
            Dict[str, Any]: Statistics about the Excel data
        """
        try:
            worksheet = workbook.active
            summary = self.get_excel_data_summary(workbook)
            
            # Basic statistics
            stats = {
                "total_videos": summary['data_rows'],
                "total_columns": summary['header_count'],
                "file_size_mb": Path(workbook.path).stat().st_size / (1024 * 1024) if hasattr(workbook, 'path') else 0,
                "last_modified": datetime.fromtimestamp(Path(workbook.path).stat().st_mtime) if hasattr(workbook, 'path') else None
            }
            
            # Try to extract some meaningful statistics from data
            try:
                # View count statistics
                view_counts = self.extract_specific_data(workbook, "View Count")
                if view_counts:
                    stats["avg_views"] = sum(view_counts) / len(view_counts)
                    stats["max_views"] = max(view_counts)
                    stats["min_views"] = min(view_counts)
                
                # Like count statistics
                like_counts = self.extract_specific_data(workbook, "Like Count")
                if like_counts:
                    stats["avg_likes"] = sum(like_counts) / len(like_counts)
                    stats["max_likes"] = max(like_counts)
                    stats["min_likes"] = min(like_counts)
                
                # Duration statistics
                durations = self.extract_specific_data(workbook, "Duration (seconds)")
                if durations:
                    stats["avg_duration"] = sum(durations) / len(durations)
                    stats["max_duration"] = max(durations)
                    stats["min_duration"] = min(durations)
                
            except Exception as e:
                logger.warning(f"Could not extract detailed statistics: {str(e)}")
            
            logger.info(f"Generated Excel statistics: {stats}")
            return stats
            
        except Exception as e:
            logger.error(f"Error generating statistics: {str(e)}")
            return {}


# Convenience functions for backward compatibility
def show_excel_data(file_path: str = None, directory: str = "downloads") -> str:
    """
    Convenience function to display Excel data.
    
    Args:
        file_path (str, optional): Specific Excel file path
        directory (str): Directory to search for Excel files
        
    Returns:
        str: Formatted display string
    """
    reader = ExcelReader()
    return reader.show_excel_data(file_path, directory)


def find_latest_excel_file(directory: str, pattern: str = "*.xlsx") -> Optional[Path]:
    """
    Convenience function to find the latest Excel file.
    
    Args:
        directory (str): Directory to search in
        pattern (str): File pattern to match
        
    Returns:
        Optional[Path]: Path to the most recent Excel file
    """
    reader = ExcelReader()
    return reader.find_latest_excel_file(directory, pattern)
