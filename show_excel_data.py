#!/usr/bin/env python3
"""
Script to display the Excel data in a readable format.
"""

import openpyxl
from pathlib import Path

def show_excel_data():
    """Display the Excel data in a readable format."""
    
    # Find the most recent Excel file
    downloads_dir = Path("downloads")
    excel_files = list(downloads_dir.glob("tiktok_videos_metadata_*.xlsx"))
    
    if not excel_files:
        print("❌ No Excel files found!")
        return
    
    # Get the most recent file
    latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
    print(f"📊 Reading Excel file: {latest_file}")
    print("=" * 80)
    
    # Load the workbook
    workbook = openpyxl.load_workbook(latest_file)
    worksheet = workbook.active
    
    # Get headers
    headers = []
    for col in range(1, worksheet.max_column + 1):
        headers.append(worksheet.cell(row=1, column=col).value)
    
    print(f"📋 Found {len(headers)} columns:")
    for i, header in enumerate(headers, 1):
        print(f"   {i:2d}. {header}")
    
    print(f"\n📊 Found {worksheet.max_row - 1} videos:")
    print("=" * 80)
    
    # Display data for each video
    for row in range(2, worksheet.max_row + 1):
        print(f"\n🎬 Video {row - 1}:")
        print("-" * 40)
        
        # Display key information
        key_fields = [
            ('Title', 2),
            ('Uploader', 4),
            ('Duration', 10),
            ('View Count', 11),
            ('Like Count', 12),
            ('Comment Count', 13),
            ('Repost Count', 14),
            ('Hashtags', 15),
            ('Upload Date', 8),
            ('Video ID', 1)
        ]
        
        for field_name, col in key_fields:
            if col <= len(headers):
                value = worksheet.cell(row=row, column=col).value
                if value:
                    # Truncate long values
                    if isinstance(value, str) and len(value) > 50:
                        value = value[:47] + "..."
                    print(f"   {field_name}: {value}")
        
        # Show full description if available
        description = worksheet.cell(row=row, column=3).value
        if description:
            print(f"   Description: {description[:100]}...")
    
    print("\n" + "=" * 80)
    print(f"✅ Excel file contains comprehensive metadata for {worksheet.max_row - 1} videos")
    print(f"📁 File location: {latest_file}")

if __name__ == "__main__":
    show_excel_data()

