"""
TikTok Video Downloader GUI

A graphical user interface for downloading TikTok videos using tkinter.
This provides an easy-to-use interface for users who prefer GUI over command line.

Usage:
    python tiktok_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import os
from pathlib import Path
from typing import Optional
import yt_dlp
from colorama import Fore, Style
import re
import openpyxl

# Import our downloader class
from tiktok_downloader import TikTokDownloader


class TikTokDownloaderGUI:
    """
    GUI class for TikTok video downloader.
    
    This class provides a user-friendly interface for downloading TikTok videos
    with features like drag-and-drop, batch processing, and real-time progress updates.
    
    Attributes:
        root (tk.Tk): Main tkinter window
        downloader (TikTokDownloader): Instance of the downloader class
        message_queue (queue.Queue): Queue for thread-safe message updates
    """
    
    def __init__(self):
        """Initialize the GUI application."""
        self.root = tk.Tk()
        self.root.title("TikTok Video Downloader")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # Initialize downloader
        self.downloader = TikTokDownloader()
        self.message_queue = queue.Queue()
        
        # Excel export variables
        self.excel_export_var = tk.BooleanVar(value=True)
        self.excel_filename_var = tk.StringVar(value="tiktok_videos_metadata.xlsx")
        
        # Configure styles
        self.setup_styles()
        
        # Create GUI elements
        self.create_widgets()
        
        # Start message processing
        self.process_messages()
    
    def setup_styles(self):
        """Configure ttk styles for a modern look."""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'))
        style.configure('Heading.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Success.TLabel', foreground='green')
        style.configure('Error.TLabel', foreground='red')
        style.configure('Info.TLabel', foreground='blue')
    
    def create_widgets(self):
        """Create and arrange all GUI widgets."""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="TikTok Video Downloader", style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # URL input section
        self.create_url_section(main_frame)
        
        # Excel import section
        self.create_excel_import_section(main_frame)
        
        # Options section
        self.create_options_section(main_frame)
        
        # Download button
        self.create_download_section(main_frame)
        
        # Progress and log section
        self.create_progress_section(main_frame)
        
        # Status bar
        self.create_status_bar(main_frame)
    
    def create_url_section(self, parent):
        """Create URL input section."""
        # URL input frame
        url_frame = ttk.LabelFrame(parent, text="Video URL", padding="10")
        url_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        url_frame.columnconfigure(1, weight=1)
        
        # URL label
        ttk.Label(url_frame, text="TikTok URL:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        # URL entry
        self.url_var = tk.StringVar()
        self.url_entry = ttk.Entry(url_frame, textvariable=self.url_var, width=60)
        self.url_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        # Paste button
        paste_btn = ttk.Button(url_frame, text="Paste", command=self.paste_url)
        paste_btn.grid(row=0, column=2, padx=(0, 10))
        
        # Clear button
        clear_btn = ttk.Button(url_frame, text="Clear", command=self.clear_url)
        clear_btn.grid(row=0, column=3)
        
        # Batch mode checkbox
        self.batch_mode_var = tk.BooleanVar()
        batch_check = ttk.Checkbutton(url_frame, text="Batch Mode (one URL per line)", 
                                    variable=self.batch_mode_var, command=self.toggle_batch_mode)
        batch_check.grid(row=1, column=0, columnspan=4, sticky=tk.W, pady=(10, 0))
        
        # Batch text area (initially hidden)
        self.batch_text = scrolledtext.ScrolledText(url_frame, height=5, width=70)
        self.batch_text.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(10, 0))
        self.batch_text.grid_remove()  # Hidden by default
    
    def create_excel_import_section(self, parent):
        """Create Excel import section."""
        # Excel import frame
        excel_frame = ttk.LabelFrame(parent, text="Excel File Import", padding="10")
        excel_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        excel_frame.columnconfigure(1, weight=1)
        
        # Excel file selection
        ttk.Label(excel_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.excel_file_var = tk.StringVar()
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel_file_var, width=50)
        excel_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        browse_excel_btn = ttk.Button(excel_frame, text="Browse", command=self.browse_excel_file)
        browse_excel_btn.grid(row=0, column=2, padx=(0, 10))
        
        load_excel_btn = ttk.Button(excel_frame, text="Load Columns", command=self.load_excel_columns)
        load_excel_btn.grid(row=0, column=3)
        
        # URL column selection
        ttk.Label(excel_frame, text="URL Column:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        self.url_column_var = tk.StringVar()
        self.url_column_combo = ttk.Combobox(excel_frame, textvariable=self.url_column_var, 
                                           state="readonly", width=20)
        self.url_column_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        # Preview button
        preview_btn = ttk.Button(excel_frame, text="Preview URLs", command=self.preview_excel_urls)
        preview_btn.grid(row=1, column=2, padx=(0, 10), pady=(10, 0))
        
        # Download from Excel button
        download_excel_btn = ttk.Button(excel_frame, text="Download from Excel", 
                                      command=self.start_excel_download, style='Accent.TButton')
        download_excel_btn.grid(row=1, column=3, pady=(10, 0))
        
        # Status label for Excel import
        self.excel_status_var = tk.StringVar(value="No Excel file selected")
        excel_status_label = ttk.Label(excel_frame, textvariable=self.excel_status_var, 
                                     font=('Arial', 9), foreground='gray')
        excel_status_label.grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=(5, 0))
    
    def create_options_section(self, parent):
        """Create download options section."""
        # Options frame
        options_frame = ttk.LabelFrame(parent, text="Download Options", padding="10")
        options_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Output directory
        ttk.Label(options_frame, text="Output Directory:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.output_dir_var = tk.StringVar(value="downloads")
        output_entry = ttk.Entry(options_frame, textvariable=self.output_dir_var, width=40)
        output_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        browse_btn = ttk.Button(options_frame, text="Browse", command=self.browse_output_dir)
        browse_btn.grid(row=0, column=2)
        
        # Quality selection
        ttk.Label(options_frame, text="Quality:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        self.quality_var = tk.StringVar(value="best")
        quality_combo = ttk.Combobox(options_frame, textvariable=self.quality_var, 
                                   values=["best", "worst", "720p", "480p", "360p"], 
                                   state="readonly", width=15)
        quality_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        # Audio only checkbox
        self.audio_only_var = tk.BooleanVar()
        audio_check = ttk.Checkbutton(options_frame, text="Audio Only", variable=self.audio_only_var)
        audio_check.grid(row=1, column=2, sticky=tk.W, pady=(10, 0))
        
        # Metadata checkbox
        self.metadata_var = tk.BooleanVar(value=False)
        metadata_check = ttk.Checkbutton(options_frame, text="Include Metadata", variable=self.metadata_var)
        metadata_check.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
        
        # Excel export checkbox
        self.excel_export_var = tk.BooleanVar(value=True)
        excel_check = ttk.Checkbutton(options_frame, text="Export to Excel", variable=self.excel_export_var)
        excel_check.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
        
        # Excel filename
        ttk.Label(options_frame, text="Excel Filename:").grid(row=4, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        self.excel_filename_var = tk.StringVar(value="tiktok_videos_metadata.xlsx")
        excel_entry = ttk.Entry(options_frame, textvariable=self.excel_filename_var, width=30)
        excel_entry.grid(row=4, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        # Custom naming field
        ttk.Label(options_frame, text="Custom Base Name:").grid(row=5, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        self.custom_name_var = tk.StringVar(value="tiktok_video")
        custom_name_entry = ttk.Entry(options_frame, textvariable=self.custom_name_var, width=30)
        custom_name_entry.grid(row=5, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        
        ttk.Label(options_frame, text="(Videos will be named: name__1, name__2, etc.)", 
                 font=('Arial', 9), foreground='gray').grid(row=5, column=2, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        
        # Process existing button
        process_existing_btn = ttk.Button(options_frame, text="Process Existing Downloads", 
                                        command=self.process_existing_downloads)
        process_existing_btn.grid(row=6, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
    
    def create_download_section(self, parent):
        """Create download control section."""
        # Download frame
        download_frame = ttk.Frame(parent)
        download_frame.grid(row=4, column=0, columnspan=3, pady=(0, 10))
        
        # Download button
        self.download_btn = ttk.Button(download_frame, text="Download Video", 
                                     command=self.start_download, style='Accent.TButton')
        self.download_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Stop button
        self.stop_btn = ttk.Button(download_frame, text="Stop", command=self.stop_download, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT)
    
    def create_progress_section(self, parent):
        """Create progress and log section."""
        # Progress frame
        progress_frame = ttk.LabelFrame(parent, text="Progress & Log", padding="10")
        progress_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        progress_frame.rowconfigure(1, weight=1)
        parent.rowconfigure(5, weight=1)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                          mode='indeterminate')
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Log text area
        self.log_text = scrolledtext.ScrolledText(progress_frame, height=15, width=80)
        self.log_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Clear log button
        clear_log_btn = ttk.Button(progress_frame, text="Clear Log", command=self.clear_log)
        clear_log_btn.grid(row=2, column=0, sticky=tk.W, pady=(10, 0))
    
    def create_status_bar(self, parent):
        """Create status bar."""
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(parent, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E))
    
    def toggle_batch_mode(self):
        """Toggle between single URL and batch mode."""
        if self.batch_mode_var.get():
            self.url_entry.grid_remove()
            self.batch_text.grid()
        else:
            self.batch_text.grid_remove()
            self.url_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
    
    def paste_url(self):
        """Paste URL from clipboard."""
        try:
            clipboard_text = self.root.clipboard_get()
            if self.batch_mode_var.get():
                self.batch_text.insert(tk.END, clipboard_text + "\n")
            else:
                self.url_var.set(clipboard_text)
        except tk.TclError:
            messagebox.showwarning("Clipboard Error", "No text in clipboard")
    
    def clear_url(self):
        """Clear URL input."""
        if self.batch_mode_var.get():
            self.batch_text.delete(1.0, tk.END)
        else:
            self.url_var.set("")
    
    def browse_output_dir(self):
        """Browse for output directory."""
        directory = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if directory:
            self.output_dir_var.set(directory)
    
    def browse_excel_file(self):
        """Browse for Excel file."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.excel_file_var.set(file_path)
            self.excel_status_var.set(f"Selected: {os.path.basename(file_path)}")
    
    def load_excel_columns(self):
        """Load column names from Excel file."""
        excel_path = self.excel_file_var.get().strip()
        if not excel_path:
            messagebox.showerror("Error", "Please select an Excel file first")
            return
        
        if not os.path.exists(excel_path):
            messagebox.showerror("Error", "Selected file does not exist")
            return
        
        try:
            # Read Excel file to get column names using openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            
            # Get column names from first row
            columns = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    columns.append(str(cell_value))
            
            wb.close()
            
            # Update combobox with column names
            self.url_column_combo['values'] = columns
            
            # Try to auto-select URL column
            url_columns = [col for col in columns if 'url' in col.lower()]
            if url_columns:
                self.url_column_var.set(url_columns[0])
                self.excel_status_var.set(f"Found {len(columns)} columns, auto-selected: {url_columns[0]}")
            else:
                self.excel_status_var.set(f"Found {len(columns)} columns, please select URL column")
            
            self.log_message(f"Loaded Excel file with {len(columns)} columns")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
            self.log_message(f"Error loading Excel file: {str(e)}", "ERROR")
    
    def preview_excel_urls(self):
        """Preview URLs from selected column."""
        excel_path = self.excel_file_var.get().strip()
        url_column = self.url_column_var.get().strip()
        
        if not excel_path or not url_column:
            messagebox.showerror("Error", "Please select Excel file and URL column first")
            return
        
        try:
            # Read Excel file using openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            
            # Find the column index for the URL column
            url_col_idx = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == url_column:
                    url_col_idx = col
                    break
            
            if url_col_idx is None:
                wb.close()
                messagebox.showerror("Error", f"Column '{url_column}' not found in Excel file")
                return
            
            # Get URLs from selected column
            urls = []
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=url_col_idx).value
                if cell_value:
                    urls.append(str(cell_value).strip())
            
            wb.close()
            
            valid_urls = self.validate_urls(urls)
            
            # Show preview dialog
            preview_text = f"Total URLs in column '{url_column}': {len(urls)}\n"
            preview_text += f"Valid TikTok URLs: {len(valid_urls)}\n\n"
            preview_text += "First 5 URLs:\n"
            
            for i, url in enumerate(valid_urls[:5], 1):
                preview_text += f"{i}. {url}\n"
            
            if len(valid_urls) > 5:
                preview_text += f"... and {len(valid_urls) - 5} more URLs"
            
            messagebox.showinfo("URL Preview", preview_text)
            self.log_message(f"Preview: {len(valid_urls)} valid URLs found in Excel")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to preview URLs: {str(e)}")
            self.log_message(f"Error previewing URLs: {str(e)}", "ERROR")
    
    def get_excel_urls(self) -> list:
        """Get URLs from Excel file."""
        excel_path = self.excel_file_var.get().strip()
        url_column = self.url_column_var.get().strip()
        
        if not excel_path or not url_column:
            return []
        
        try:
            # Read Excel file using openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            
            # Find the column index for the URL column
            url_col_idx = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == url_column:
                    url_col_idx = col
                    break
            
            if url_col_idx is None:
                wb.close()
                self.log_message(f"Column '{url_column}' not found in Excel file", "ERROR")
                return []
            
            # Get URLs from selected column
            urls = []
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=url_col_idx).value
                if cell_value:
                    urls.append(str(cell_value).strip())
            
            wb.close()
            
            valid_urls = self.validate_urls(urls)
            
            return valid_urls
            
        except Exception as e:
            self.log_message(f"Error reading Excel file: {str(e)}", "ERROR")
            return []
    
    def start_excel_download(self):
        """Start download process from Excel file."""
        urls = self.get_excel_urls()
        
        if not urls:
            messagebox.showerror("Error", "No valid TikTok URLs found in Excel file")
            return
        
        # Check for duplicate URLs
        unique_urls = []
        duplicate_urls = []
        for url in urls:
            if url in unique_urls:
                duplicate_urls.append(url)
            else:
                unique_urls.append(url)
        
        if duplicate_urls:
            self.log_message(f"Removed {len(duplicate_urls)} duplicate URLs from Excel", "WARNING")
        
        if not unique_urls:
            messagebox.showerror("Error", "No unique URLs to download from Excel")
            return
        
        # Update UI state
        self.download_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress_bar.start()
        self.status_var.set("Downloading from Excel...")
        
        # Start download thread
        self.download_thread = threading.Thread(target=self.download_worker, args=(unique_urls,))
        self.download_thread.daemon = True
        self.download_thread.start()
    
    def clear_log(self):
        """Clear log text area."""
        self.log_text.delete(1.0, tk.END)
    
    def log_message(self, message: str, level: str = "INFO"):
        """Add message to log with timestamp and level."""
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {level}: {message}\n"
        
        # Add to queue for thread-safe update
        self.message_queue.put(formatted_message)
    
    def process_messages(self):
        """Process messages from queue (thread-safe)."""
        try:
            while True:
                message = self.message_queue.get_nowait()
                self.log_text.insert(tk.END, message)
                self.log_text.see(tk.END)
                self.log_text.update_idletasks()
        except queue.Empty:
            pass
        
        # Schedule next check
        self.root.after(100, self.process_messages)
    
    def validate_urls(self, urls: list) -> list:
        """Validate and filter URLs."""
        valid_urls = []
        for url in urls:
            url = url.strip()
            if url and self.downloader.validate_url(url):
                valid_urls.append(url)
            elif url:
                self.log_message(f"Invalid URL: {url}", "WARNING")
        
        return valid_urls
    
    def get_urls(self) -> list:
        """Get URLs from input (single or batch)."""
        if self.batch_mode_var.get():
            text = self.batch_text.get(1.0, tk.END)
            urls = text.strip().split('\n')
        else:
            url = self.url_var.get().strip()
            urls = [url] if url else []
        
        return self.validate_urls(urls)
    
    def start_download(self):
        """Start download process in separate thread."""
        urls = self.get_urls()
        
        if not urls:
            messagebox.showerror("Error", "Please enter valid TikTok URL(s)")
            return
        
        # Check for duplicate URLs
        unique_urls = []
        duplicate_urls = []
        for url in urls:
            if url in unique_urls:
                duplicate_urls.append(url)
            else:
                unique_urls.append(url)
        
        if duplicate_urls:
            self.log_message(f"Removed {len(duplicate_urls)} duplicate URLs", "WARNING")
        
        if not unique_urls:
            messagebox.showerror("Error", "No unique URLs to download")
            return
        
        # Update UI state
        self.download_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress_bar.start()
        self.status_var.set("Downloading...")
        
        # Start download thread
        self.download_thread = threading.Thread(target=self.download_worker, args=(unique_urls,))
        self.download_thread.daemon = True
        self.download_thread.start()
    
    def stop_download(self):
        """Stop download process."""
        self.status_var.set("Stopping...")
        # Note: yt-dlp doesn't have a built-in stop mechanism
        # This is a placeholder for future implementation
        self.finish_download()
    
    def download_worker(self, urls: list):
        """Worker thread for downloading videos."""
        try:
            # Update downloader settings
            self.downloader.output_dir = Path(self.output_dir_var.get())
            self.downloader.quality = self.quality_var.get()
            self.downloader.extract_audio = self.audio_only_var.get()
            self.downloader.add_metadata = self.metadata_var.get()
            
            # Reset video counter for new batch
            self.downloader.reset_video_counter()
            
            self.downloader.ydl_opts = self.downloader._configure_ydl_options()
            
            # Set custom naming and Excel export settings
            custom_name = self.custom_name_var.get().strip()
            if custom_name:
                self.downloader.custom_base_name = custom_name
                self.log_message(f"Using custom naming: {custom_name}__1, {custom_name}__2, etc.")
                
                # Reconfigure ydl_opts with custom naming
                self.downloader.ydl_opts = self.downloader._configure_ydl_options()
            
            # Set Excel export settings
            if self.excel_export_var.get():
                # Use custom name for Excel file with __excel suffix
                if custom_name:
                    excel_filename = f"{custom_name}__excel.xlsx"
                else:
                    excel_filename = self.excel_filename_var.get()
                excel_path = self.downloader.output_dir / excel_filename
                self.downloader.excel_file = excel_path
                self.log_message("Excel export enabled")
                
                # Clear any existing data in the Excel file for fresh batch
                self.downloader.workbook = openpyxl.Workbook()
                self.downloader.worksheet = self.downloader.workbook.active
                self.downloader.worksheet.title = "TikTok Videos Metadata"
                self.downloader._setup_excel_headers()
            
            # Create output directory
            self.downloader.output_dir.mkdir(exist_ok=True)
            
            self.log_message(f"Starting download of {len(urls)} video(s)")
            self.log_message(f"Output directory: {self.downloader.output_dir}")
            self.log_message(f"Quality: {self.downloader.quality}")
            
            # Download videos
            for i, url in enumerate(urls, 1):
                self.log_message(f"Processing video {i}/{len(urls)}: {url}")
                
                try:
                    # Check if URL is already downloaded before attempting download
                    if self.downloader.is_url_already_downloaded(url):
                        self.log_message(f"Video {i} already downloaded, skipping", "INFO")
                        continue
                    
                    self.log_message(f"Starting download for video {i}...", "INFO")
                    success = self.downloader.download_video(url)
                    
                    if success:
                        self.log_message(f"Successfully downloaded video {i}", "SUCCESS")
                    else:
                        self.log_message(f"Failed to download video {i}", "ERROR")
                        
                except Exception as e:
                    self.log_message(f"Error downloading video {i}: {str(e)}", "ERROR")
            
            # Save Excel file if enabled
            if self.excel_export_var.get():
                self.log_message("Saving Excel file with metadata...")
                try:
                    self.downloader.save_excel_file()
                    self.log_message(f"Excel file saved: {self.downloader.excel_file}", "SUCCESS")
                except Exception as e:
                    self.log_message(f"Error saving Excel file: {str(e)}", "ERROR")
            
            self.log_message("Download process completed", "SUCCESS")
            
        except Exception as e:
            self.log_message(f"Unexpected error: {str(e)}", "ERROR")
        
        finally:
            # Update UI on main thread
            self.root.after(0, self.finish_download)
    
    def finish_download(self):
        """Finish download process and update UI."""
        self.download_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.progress_bar.stop()
        self.status_var.set("Ready")
    
    def process_existing_downloads(self):
        """Process existing downloads and create Excel file."""
        try:
            # Update downloader settings
            self.downloader.output_dir = Path(self.output_dir_var.get())
            
            # Set custom naming and Excel filename
            custom_name = self.custom_name_var.get().strip()
            if custom_name:
                self.downloader.custom_base_name = custom_name
                excel_filename = f"{custom_name}__excel.xlsx"
                excel_path = self.downloader.output_dir / excel_filename
                self.downloader.excel_file = excel_path
            elif self.excel_filename_var.get():
                excel_path = self.downloader.output_dir / self.excel_filename_var.get()
                self.downloader.excel_file = excel_path
            
            self.log_message("Processing existing downloads for Excel export...")
            self.status_var.set("Processing existing downloads...")
            
            # Process in separate thread
            process_thread = threading.Thread(target=self.process_existing_worker)
            process_thread.daemon = True
            process_thread.start()
            
        except Exception as e:
            self.log_message(f"Error starting process: {str(e)}", "ERROR")
    
    def process_existing_worker(self):
        """Worker thread for processing existing downloads."""
        try:
            self.downloader.process_existing_downloads()
            self.log_message("Successfully processed existing downloads", "SUCCESS")
            self.log_message(f"Excel file saved: {self.downloader.excel_file}", "SUCCESS")
            
        except Exception as e:
            self.log_message(f"Error processing existing downloads: {str(e)}", "ERROR")
        
        finally:
            # Update UI on main thread
            self.root.after(0, lambda: self.status_var.set("Ready"))
    
    def run(self):
        """Start the GUI application."""
        self.root.mainloop()


def main():
    """Main function to start the GUI application."""
    try:
        app = TikTokDownloaderGUI()
        app.run()
    except Exception as e:
        print(f"Error starting GUI: {e}")
        messagebox.showerror("Error", f"Failed to start application: {e}")


if __name__ == "__main__":
    main()
