"""
TikTok Video Downloader Module

This module provides functionality to download TikTok videos using yt-dlp.
It includes both command-line and programmatic interfaces for downloading videos.

Usage:
    # Command line
    python tiktok_downloader.py --url "https://www.tiktok.com/@user/video/1234567890"
    
    # Programmatic
    from tiktok_downloader import TikTokDownloader
    downloader = TikTokDownloader()
    downloader.download_video("https://www.tiktok.com/@user/video/1234567890")
"""

import os
import sys
import logging
import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List
import yt_dlp
from colorama import init, Fore, Style
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('tiktok_downloader.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class TikTokDownloader:
    """
    A class to handle TikTok video downloads using yt-dlp.
    
    This class provides methods to download TikTok videos with various options
    including quality selection, output directory specification, and metadata extraction.
    
    Attributes:
        output_dir (str): Directory where downloaded videos will be saved
        quality (str): Preferred video quality ('best', 'worst', '720p', etc.)
        extract_audio (bool): Whether to extract audio only
        add_metadata (bool): Whether to add metadata to downloaded files
        excel_file (str): Path to the Excel file for metadata export
    """
    
    def __init__(self, output_dir: str = "downloads", quality: str = "best", 
                 extract_audio: bool = False, add_metadata: bool = True, 
                 excel_file: str = None):
        """
        Initialize the TikTok downloader with specified options.
        
        Args:
            output_dir (str): Directory to save downloaded videos (default: "downloads")
            quality (str): Preferred video quality (default: "best")
            extract_audio (bool): Extract audio only if True (default: False)
            add_metadata (bool): Add metadata to downloaded files (default: True)
            excel_file (str): Path to Excel file for metadata export (default: auto-generated)
        """
        self.output_dir = Path(output_dir)
        self.quality = quality
        self.extract_audio = extract_audio
        self.add_metadata = add_metadata
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)
        
        # Set Excel file path
        if excel_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.excel_file = self.output_dir / f"tiktok_videos_metadata_{timestamp}.xlsx"
        else:
            self.excel_file = Path(excel_file)
        
        # Initialize Excel workbook
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "TikTok Videos Metadata"
        
        # Define headers for Excel
        self.headers = [
            'Video ID', 'Title', 'Description', 'Uploader', 'Uploader ID', 
            'Channel', 'Channel ID', 'Upload Date', 'Duration (seconds)', 
            'Duration (formatted)', 'View Count', 'Like Count', 'Comment Count', 
            'Repost Count', 'Hashtags', 'Original URL', 'Thumbnail URL',
            'Video Quality', 'File Size (bytes)', 'Resolution', 'Format',
            'Download Date', 'Download Path'
        ]
        
        # Setup Excel headers
        self._setup_excel_headers()
        
        # Configure yt-dlp options
        self.ydl_opts = self._configure_ydl_options()
    
    def _setup_excel_headers(self):
        """Setup Excel worksheet with headers and formatting."""
        # Add headers
        for col, header in enumerate(self.headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col in range(1, len(self.headers) + 1):
            column_letter = get_column_letter(col)
            self.worksheet.column_dimensions[column_letter].width = 15
    
    def _extract_hashtags(self, description: str) -> str:
        """
        Extract hashtags from video description.
        
        Args:
            description (str): Video description text
            
        Returns:
            str: Comma-separated hashtags
        """
        if not description:
            return ""
        
        hashtags = []
        words = description.split()
        for word in words:
            if word.startswith('#'):
                hashtags.append(word)
        
        return ', '.join(hashtags)
    
    def _format_duration(self, duration: int) -> str:
        """
        Format duration from seconds to MM:SS format.
        
        Args:
            duration (int): Duration in seconds
            
        Returns:
            str: Formatted duration string
        """
        if not duration:
            return ""
        
        minutes = duration // 60
        seconds = duration % 60
        return f"{minutes:02d}:{seconds:02d}"
    
    def _extract_metadata_for_excel(self, info: Dict[str, Any], download_path: str = "") -> List[Any]:
        """
        Extract metadata from video info for Excel export.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            download_path (str): Path where video was downloaded
            
        Returns:
            List[Any]: List of metadata values for Excel row
        """
        # Extract hashtags from description
        description = info.get('description', '')
        hashtags = self._extract_hashtags(description)
        
        # Format duration
        duration = info.get('duration', 0)
        duration_formatted = self._format_duration(duration)
        
        # Get video format info
        format_info = info.get('format', '')
        if isinstance(format_info, str):
            video_quality = format_info
        else:
            video_quality = format_info.get('format_note', '') or format_info.get('format', '')
        
        # Get file size
        filesize = info.get('filesize', 0)
        if not filesize and 'format' in info:
            filesize = info['format'].get('filesize', 0)
        
        # Get resolution
        width = info.get('width', 0)
        height = info.get('height', 0)
        resolution = f"{width}x{height}" if width and height else ""
        
        # Get thumbnail URL
        thumbnail = info.get('thumbnail', '')
        
        # Format upload date
        upload_date = info.get('upload_date', '')
        if upload_date:
            try:
                formatted_date = f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:8]}"
            except:
                formatted_date = upload_date
        else:
            formatted_date = ""
        
        # Current download date
        download_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        return [
            info.get('id', ''),
            info.get('title', ''),
            description,
            info.get('uploader', ''),
            info.get('uploader_id', ''),
            info.get('channel', ''),
            info.get('channel_id', ''),
            formatted_date,
            duration,
            duration_formatted,
            info.get('view_count', 0),
            info.get('like_count', 0),
            info.get('comment_count', 0),
            info.get('repost_count', 0),
            hashtags,
            info.get('webpage_url', ''),
            thumbnail,
            video_quality,
            filesize,
            resolution,
            info.get('ext', ''),
            download_date,
            download_path
        ]
    
    def _add_metadata_to_excel(self, info: Dict[str, Any], download_path: str = ""):
        """
        Add video metadata to Excel worksheet.
        
        Args:
            info (Dict[str, Any]): Video information dictionary
            download_path (str): Path where video was downloaded
        """
        try:
            # Get next row number
            next_row = self.worksheet.max_row + 1
            
            # Extract metadata
            metadata = self._extract_metadata_for_excel(info, download_path)
            
            # Add data to worksheet
            for col, value in enumerate(metadata, 1):
                cell = self.worksheet.cell(row=next_row, column=col, value=value)
                
                # Format numbers
                if col in [11, 12, 13, 14, 19]:  # View count, like count, comment count, repost count, file size
                    if value and value != 0:
                        cell.number_format = '#,##0'
            
            # Auto-adjust column widths
            for col in range(1, len(self.headers) + 1):
                column_letter = get_column_letter(col)
                current_width = self.worksheet.column_dimensions[column_letter].width
                cell_value = self.worksheet.cell(row=next_row, column=col).value
                if cell_value:
                    content_length = len(str(cell_value))
                    if content_length > current_width:
                        self.worksheet.column_dimensions[column_letter].width = min(content_length + 2, 50)
            
            logger.info(f"Added metadata to Excel: {info.get('title', 'Unknown')}")
            
        except Exception as e:
            logger.error(f"Error adding metadata to Excel: {e}")
    
    def save_excel_file(self):
        """Save the Excel file with all collected metadata."""
        try:
            self.workbook.save(str(self.excel_file))
            print(f"{Fore.GREEN}Excel file saved: {self.excel_file}{Style.RESET_ALL}")
            logger.info(f"Excel file saved: {self.excel_file}")
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            print(f"{Fore.RED}Error saving Excel file: {e}{Style.RESET_ALL}")
    
    def process_existing_downloads(self):
        """
        Process existing downloaded videos and create Excel file from their metadata.
        This method reads existing .info.json files and adds them to the Excel.
        """
        print(f"{Fore.CYAN}Processing existing downloads for Excel export...{Style.RESET_ALL}")
        
        # Find all .info.json files in the output directory
        info_files = list(self.output_dir.glob("*.info.json"))
        
        if not info_files:
            print(f"{Fore.YELLOW}No existing .info.json files found in {self.output_dir}{Style.RESET_ALL}")
            return
        
        print(f"{Fore.GREEN}Found {len(info_files)} existing video metadata files{Style.RESET_ALL}")
        
        processed_count = 0
        for info_file in info_files:
            try:
                # Read the JSON file
                with open(info_file, 'r', encoding='utf-8') as f:
                    info = json.load(f)
                
                # Find corresponding video file
                video_path = ""
                title = info.get('title', '')
                ext = info.get('ext', 'mp4')
                
                # Look for video file with similar name
                for video_file in self.output_dir.glob(f"*.{ext}"):
                    if title.lower() in video_file.name.lower():
                        video_path = str(video_file)
                        break
                
                # Add to Excel
                self._add_metadata_to_excel(info, video_path)
                processed_count += 1
                
                print(f"{Fore.GREEN}Processed: {title[:50]}...{Style.RESET_ALL}")
                
            except Exception as e:
                logger.error(f"Error processing {info_file}: {e}")
                print(f"{Fore.RED}Error processing {info_file.name}: {e}{Style.RESET_ALL}")
        
        # Save the Excel file
        self.save_excel_file()
        print(f"{Fore.GREEN}Successfully processed {processed_count} videos{Style.RESET_ALL}")
    
    def _configure_ydl_options(self) -> Dict[str, Any]:
        """
        Configure yt-dlp options based on user preferences.
        
        Returns:
            Dict[str, Any]: Configured yt-dlp options dictionary
        """
        options = {
            'outtmpl': str(self.output_dir / '%(title)s.%(ext)s'),
            'format': 'bestaudio/best' if self.extract_audio else self.quality,
            'writesubtitles': False,
            'writeautomaticsub': False,
            'ignoreerrors': False,
            'no_warnings': False,
            'quiet': False,
            'verbose': True,
        }
        
        # Add metadata options if enabled
        if self.add_metadata:
            options.update({
                'writethumbnail': True,
                'writedescription': True,
                'writeinfojson': True,
            })
        
        return options
    
    def validate_url(self, url: str) -> bool:
        """
        Validate if the provided URL is a valid TikTok URL.
        
        Args:
            url (str): The URL to validate
            
        Returns:
            bool: True if URL is valid, False otherwise
        """
        valid_domains = [
            'tiktok.com',
            'vm.tiktok.com',
            'vt.tiktok.com',
            'www.tiktok.com'
        ]
        
        try:
            import urllib.parse
            parsed = urllib.parse.urlparse(url)
            return any(domain in parsed.netloc for domain in valid_domains)
        except Exception:
            return False
    
    def get_video_info(self, url: str) -> Optional[Dict[str, Any]]:
        """
        Extract video information without downloading.
        
        Args:
            url (str): TikTok video URL
            
        Returns:
            Optional[Dict[str, Any]]: Video information dictionary or None if failed
        """
        try:
            with yt_dlp.YoutubeDL(self.ydl_opts) as ydl:
                logger.info(f"Extracting info for: {url}")
                info = ydl.extract_info(url, download=False)
                return info
        except Exception as e:
            logger.error(f"Failed to extract video info: {e}")
            return None
    
    def download_video(self, url: str) -> bool:
        """
        Download a TikTok video from the provided URL.
        
        Args:
            url (str): TikTok video URL to download
            
        Returns:
            bool: True if download successful, False otherwise
        """
        if not self.validate_url(url):
            logger.error(f"Invalid TikTok URL: {url}")
            print(f"{Fore.RED}Error: Invalid TikTok URL provided{Style.RESET_ALL}")
            return False
        
        try:
            print(f"{Fore.CYAN}Starting download for: {url}{Style.RESET_ALL}")
            logger.info(f"Starting download for URL: {url}")
            
            with yt_dlp.YoutubeDL(self.ydl_opts) as ydl:
                # Extract info first
                info = ydl.extract_info(url, download=False)
                if not info:
                    logger.error("Failed to extract video information")
                    return False
                
                print(f"{Fore.GREEN}Video Title: {info.get('title', 'Unknown')}{Style.RESET_ALL}")
                print(f"{Fore.GREEN}Duration: {info.get('duration', 'Unknown')} seconds{Style.RESET_ALL}")
                print(f"{Fore.GREEN}Uploader: {info.get('uploader', 'Unknown')}{Style.RESET_ALL}")
                
                # Download the video
                ydl.download([url])
                
                # Get the downloaded file path
                download_path = ""
                if info.get('title'):
                    # Try to find the downloaded file
                    title = info.get('title', '')
                    ext = info.get('ext', 'mp4')
                    possible_filename = f"{title}.{ext}"
                    download_path = str(self.output_dir / possible_filename)
                    
                    # Check if file exists
                    if not os.path.exists(download_path):
                        # Try to find any file with similar name
                        for file in self.output_dir.glob(f"*{ext}"):
                            if title.lower() in file.name.lower():
                                download_path = str(file)
                                break
                
                # Add metadata to Excel
                self._add_metadata_to_excel(info, download_path)
                
                print(f"{Fore.GREEN}Download completed successfully!{Style.RESET_ALL}")
                logger.info("Download completed successfully")
                return True
                
        except yt_dlp.DownloadError as e:
            logger.error(f"Download error: {e}")
            print(f"{Fore.RED}Download failed: {e}{Style.RESET_ALL}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error during download: {e}")
            print(f"{Fore.RED}Unexpected error: {e}{Style.RESET_ALL}")
            return False
    
    def download_multiple_videos(self, urls: list) -> Dict[str, bool]:
        """
        Download multiple TikTok videos from a list of URLs.
        
        Args:
            urls (list): List of TikTok video URLs
            
        Returns:
            Dict[str, bool]: Dictionary mapping URLs to download success status
        """
        results = {}
        
        print(f"{Fore.YELLOW}Starting batch download of {len(urls)} videos...{Style.RESET_ALL}")
        
        for i, url in enumerate(urls, 1):
            print(f"\n{Fore.CYAN}[{i}/{len(urls)}] Processing: {url}{Style.RESET_ALL}")
            results[url] = self.download_video(url)
        
        # Save Excel file with all metadata
        self.save_excel_file()
        
        # Print summary
        successful = sum(results.values())
        print(f"\n{Fore.GREEN}Batch download completed!{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Successful: {successful}/{len(urls)}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Metadata saved to Excel: {self.excel_file}{Style.RESET_ALL}")
        
        return results


def main():
    """
    Main function to handle command-line interface.
    
    This function parses command-line arguments and initiates the download process.
    It supports single video downloads, batch downloads from a file, and various options.
    """
    parser = argparse.ArgumentParser(
        description="Download TikTok videos using yt-dlp",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tiktok_downloader.py --url "https://www.tiktok.com/@user/video/1234567890"
  python tiktok_downloader.py --file urls.txt --quality 720p
  python tiktok_downloader.py --url "https://tiktok.com/@user/video/1234567890" --audio-only
  python tiktok_downloader.py --process-existing
        """
    )
    
    # URL input options
    url_group = parser.add_mutually_exclusive_group(required=True)
    url_group.add_argument(
        '--url', 
        type=str, 
        help='Single TikTok video URL to download'
    )
    url_group.add_argument(
        '--file', 
        type=str, 
        help='Text file containing TikTok URLs (one per line)'
    )
    url_group.add_argument(
        '--process-existing',
        action='store_true',
        help='Process existing downloaded videos and create Excel file from metadata'
    )
    
    # Download options
    parser.add_argument(
        '--output-dir', 
        type=str, 
        default='downloads',
        help='Output directory for downloaded videos (default: downloads)'
    )
    parser.add_argument(
        '--quality', 
        type=str, 
        default='best',
        help='Video quality preference (default: best)'
    )
    parser.add_argument(
        '--audio-only', 
        action='store_true',
        help='Download audio only (no video)'
    )
    parser.add_argument(
        '--no-metadata', 
        action='store_true',
        help='Skip metadata extraction'
    )
    
    args = parser.parse_args()
    
    # Initialize downloader
    downloader = TikTokDownloader(
        output_dir=args.output_dir,
        quality=args.quality,
        extract_audio=args.audio_only,
        add_metadata=not args.no_metadata
    )
    
    try:
        if args.url:
            # Single video download
            success = downloader.download_video(args.url)
            if success:
                # Save Excel file for single download
                downloader.save_excel_file()
            sys.exit(0 if success else 1)
        
        elif args.file:
            # Batch download from file
            if not os.path.exists(args.file):
                print(f"{Fore.RED}Error: File '{args.file}' not found{Style.RESET_ALL}")
                sys.exit(1)
            
            with open(args.file, 'r') as f:
                urls = [line.strip() for line in f if line.strip()]
            
            if not urls:
                print(f"{Fore.RED}Error: No valid URLs found in file{Style.RESET_ALL}")
                sys.exit(1)
            
            results = downloader.download_multiple_videos(urls)
            successful = sum(results.values())
            sys.exit(0 if successful == len(urls) else 1)
        
        elif args.process_existing:
            # Process existing downloads
            downloader.process_existing_downloads()
            sys.exit(0)
    
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}Download interrupted by user{Style.RESET_ALL}")
        sys.exit(1)
    except Exception as e:
        print(f"{Fore.RED}Unexpected error: {e}{Style.RESET_ALL}")
        sys.exit(1)


if __name__ == "__main__":
    main()
