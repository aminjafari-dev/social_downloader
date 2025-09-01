#!/usr/bin/env python3
"""
Test script to debug the download issue by simulating the download process.
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloader'))

from tiktok_downloader import TikTokDownloader
from pathlib import Path

def test_download_process():
    """Test the download process step by step."""
    
    print("🧪 Testing Download Process")
    print("=" * 50)
    
    # Initialize downloader
    downloader = TikTokDownloader(
        output_dir="downloads",
        quality="best",
        extract_audio=False,
        add_metadata=True
    )
    
    # Set custom base name
    downloader.custom_base_name = "test_video"
    
    # Test URLs (these are example URLs - replace with real ones for testing)
    test_urls = [
        "https://www.tiktok.com/@user1/video/1234567890123456789",
        "https://www.tiktok.com/@user2/video/9876543210987654321",
        "https://www.tiktok.com/@user3/video/5556667778889990001"
    ]
    
    print(f"📊 Testing with {len(test_urls)} URLs")
    print(f"🎯 Custom base name: {downloader.custom_base_name}")
    print(f"📁 Output directory: {downloader.output_dir}")
    
    # Reset video counter
    print("\n🔄 Resetting video counter...")
    downloader.reset_video_counter()
    print(f"📊 Video counter after reset: {downloader.video_counter}")
    
    # Test each URL
    for i, url in enumerate(test_urls, 1):
        print(f"\n🎬 Processing URL {i}: {url}")
        
        # Check if already downloaded
        already_downloaded = downloader.is_url_already_downloaded(url)
        print(f"📋 Already downloaded: {already_downloaded}")
        
        if not already_downloaded:
            print(f"📊 Video counter before download: {downloader.video_counter}")
            
            # Try to download (this will fail with fake URLs, but we can see the counter logic)
            try:
                # Just simulate the process without actually downloading
                print("🔄 Simulating download process...")
                
                # Check what the filename would be
                filename = downloader._get_custom_filename()
                print(f"📝 Generated filename: {filename}")
                print(f"📊 Video counter after filename generation: {downloader.video_counter}")
                
            except Exception as e:
                print(f"⚠️  Expected error with fake URL: {e}")
        else:
            print("⏭️  Skipping already downloaded URL")
    
    print(f"\n📊 Final video counter: {downloader.video_counter}")
    
    # Check Excel file if it exists
    if downloader.excel_file.exists():
        print(f"\n📊 Excel file exists: {downloader.excel_file}")
        print("📋 Checking Excel file content...")
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(downloader.excel_file))
            ws = wb.active
            
            print(f"📊 Excel has {ws.max_row} rows (including header)")
            
            if ws.max_row > 1:
                print("📋 Data rows:")
                for row in range(2, ws.max_row + 1):
                    url = ws.cell(row=row, column=16).value  # Original URL column
                    path = ws.cell(row=row, column=len(downloader.headers)).value  # Download Path column
                    print(f"  Row {row}: URL={url}, Path={path}")
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
    
    print("\n🎉 Download process test completed!")

if __name__ == "__main__":
    test_download_process()
