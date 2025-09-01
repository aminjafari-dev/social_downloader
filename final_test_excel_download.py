#!/usr/bin/env python3
"""
Final comprehensive test for Excel download feature.
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloader'))

from tiktok_downloader import TikTokDownloader
from pathlib import Path
import openpyxl

def final_test_excel_download():
    """Final comprehensive test for Excel download feature."""
    
    print("🎯 Final Excel Download Test")
    print("=" * 50)
    
    # Initialize downloader
    downloader = TikTokDownloader(
        output_dir="downloads",
        quality="best",
        extract_audio=False,
        add_metadata=True
    )
    
    # Set custom base name
    downloader.custom_base_name = "final_test"
    
    # Test URLs (replace with real TikTok URLs for actual testing)
    test_urls = [
        "https://www.tiktok.com/@user1/video/1234567890123456789",
        "https://www.tiktok.com/@user2/video/9876543210987654321",
        "https://www.tiktok.com/@user3/video/5556667778889990001"
    ]
    
    print(f"📊 Testing with {len(test_urls)} URLs")
    print(f"🎯 Custom base name: {downloader.custom_base_name}")
    print(f"📁 Output directory: {downloader.output_dir}")
    print(f"📄 Excel file: {downloader.excel_file}")
    
    # Reset video counter
    print("\n🔄 Resetting video counter...")
    downloader.reset_video_counter()
    print(f"📊 Video counter after reset: {downloader.video_counter}")
    
    # Test the complete download flow
    successful_downloads = 0
    failed_downloads = 0
    skipped_downloads = 0
    
    for i, url in enumerate(test_urls, 1):
        print(f"\n🎬 Processing URL {i}: {url}")
        print("-" * 40)
        
        # Check if already downloaded
        already_downloaded = downloader.is_url_already_downloaded(url)
        print(f"📋 Already downloaded: {already_downloaded}")
        
        if already_downloaded:
            print("⏭️  Skipping already downloaded URL")
            skipped_downloads += 1
            continue
        
        # Validate URL
        is_valid = downloader.validate_url(url)
        print(f"✅ URL valid: {is_valid}")
        
        if not is_valid:
            print("❌ Invalid URL, skipping")
            failed_downloads += 1
            continue
        
        # Simulate download (since these are fake URLs)
        print("🔄 Simulating download...")
        
        # Create mock video info
        mock_info = {
            'id': f'video_{i}',
            'title': f'Final Test Video {i}',
            'description': f'Description for final test video {i}',
            'uploader': f'user{i}',
            'uploader_id': f'user{i}_id',
            'channel': f'channel{i}',
            'channel_id': f'channel{i}_id',
            'upload_date': '20250101',
            'duration': 30,
            'view_count': 1000,
            'like_count': 100,
            'comment_count': 50,
            'repost_count': 10,
            'ext': 'mp4',
            'filesize': 1024000,
            'width': 1920,
            'height': 1080,
            'format': 'mp4',
            'webpage_url': url
        }
        
        # Generate download path
        ext = mock_info.get('ext', 'mp4')
        current_counter = downloader.video_counter
        download_path = str(downloader.output_dir / f"{downloader.custom_base_name}__{current_counter}.{ext}")
        
        print(f"📝 Generated download path: {download_path}")
        
        # Add metadata to Excel (this simulates successful download)
        downloader._add_metadata_to_excel(mock_info, download_path)
        print("✅ Added metadata to Excel")
        
        # Increment video counter
        downloader.video_counter += 1
        print(f"📊 Video counter after processing: {downloader.video_counter}")
        
        successful_downloads += 1
    
    print(f"\n📊 Download Summary:")
    print(f"✅ Successful: {successful_downloads}")
    print(f"❌ Failed: {failed_downloads}")
    print(f"⏭️  Skipped: {skipped_downloads}")
    print(f"📊 Total processed: {successful_downloads + failed_downloads + skipped_downloads}")
    
    # Verify Excel file
    if downloader.excel_file.exists():
        print(f"\n✅ Excel file created successfully: {downloader.excel_file}")
        
        try:
            wb = openpyxl.load_workbook(str(downloader.excel_file))
            ws = wb.active
            
            print(f"📊 Excel has {ws.max_row} rows (including header)")
            
            if ws.max_row > 1:
                print("📋 Data rows:")
                for row in range(2, ws.max_row + 1):
                    url = ws.cell(row=row, column=16).value  # Original URL column
                    path = ws.cell(row=row, column=len(downloader.headers)).value  # Download Path column
                    print(f"  Row {row}: URL={url}")
                    print(f"         Path={path}")
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
    else:
        print(f"\n❌ Excel file not created: {downloader.excel_file}")
    
    print(f"\n🎉 Final test completed!")
    print(f"📊 Final video counter: {downloader.video_counter}")
    
    if successful_downloads == len(test_urls):
        print("✅ All URLs processed successfully!")
    else:
        print(f"⚠️  Only {successful_downloads}/{len(test_urls)} URLs processed successfully")

if __name__ == "__main__":
    final_test_excel_download()
