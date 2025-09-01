#!/usr/bin/env python3
"""
Test script to create a sample Excel file with TikTok URLs for testing the Excel import feature.
Uses openpyxl directly to avoid NumPy version conflicts.
"""

import openpyxl
import os

def create_test_excel_file():
    """Create a sample Excel file with TikTok URLs for testing."""
    
    # Sample TikTok URLs (these are example URLs - replace with real ones for testing)
    sample_urls = [
        "https://www.tiktok.com/@user1/video/1234567890123456789",
        "https://www.tiktok.com/@user2/video/9876543210987654321",
        "https://www.tiktok.com/@user3/video/5556667778889990001",
        "https://www.tiktok.com/@user4/video/1112223334445556667",
        "https://www.tiktok.com/@user5/video/9998887776665554443"
    ]
    
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TikTok URLs"
    
    # Add headers
    headers = ['Title', 'URL', 'Description', 'Category']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    titles = ['Video 1', 'Video 2', 'Video 3', 'Video 4', 'Video 5']
    descriptions = ['First video', 'Second video', 'Third video', 'Fourth video', 'Fifth video']
    categories = ['Funny', 'Dance', 'Music', 'Comedy', 'Tutorial']
    
    for row, (title, url, desc, cat) in enumerate(zip(titles, sample_urls, descriptions, categories), 2):
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=url)
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value=cat)
    
    # Create test Excel file
    excel_filename = "test_tiktok_urls.xlsx"
    wb.save(excel_filename)
    
    print(f"✅ Created test Excel file: {excel_filename}")
    print(f"📊 File contains {len(sample_urls)} sample TikTok URLs")
    print(f"📋 Columns: {headers}")
    print("\n📝 Instructions for testing:")
    print("1. Run the GUI: python main.py")
    print("2. In the 'Excel File Import' section:")
    print("   - Click 'Browse' and select this file")
    print("   - Click 'Load Columns' to load column names")
    print("   - Select 'URL' column from dropdown")
    print("   - Click 'Preview URLs' to see the URLs")
    print("   - Click 'Download from Excel' to start downloading")
    
    return excel_filename

if __name__ == "__main__":
    create_test_excel_file()

